from . import models
from openpyxl import load_workbook
from .models import FundosImobiliarios
from ..database import SessionLocal
import time

db = SessionLocal()


def import_fundos_imobiliarios():
    st = time.time()
    workbook = load_workbook("src/importacoes/rendavariavel.xlsx")
    sheet = workbook["FIIS"]
    row_count = sheet.max_row
    list_fii = []
    for row in range(2, row_count + 1):
        fii = FundosImobiliarios(
            sheet.cell(row=row, column=1).value,
            sheet.cell(row=row, column=2).value,
            sheet.cell(row=row, column=3).value,
            sheet.cell(row=row, column=4).value,
            sheet.cell(row=row, column=5).value,
            sheet.cell(row=row, column=6).value,
            sheet.cell(row=row, column=7).value,
            sheet.cell(row=row, column=8).value,
            sheet.cell(row=row, column=9).value,
            sheet.cell(row=row, column=10).value,
            sheet.cell(row=row, column=11).value,
            sheet.cell(row=row, column=12).value,
            sheet.cell(row=row, column=13).value,
            sheet.cell(row=row, column=14).value,
            sheet.cell(row=row, column=15).value,
            sheet.cell(row=row, column=16).value,
            sheet.cell(row=row, column=17).value,
            sheet.cell(row=row, column=18).value,
        )
        list_fii.append(fii)

    count = db.query(models.FundosImobiliarios).count()
    if count == 0:
        db.add_all(list_fii)
        db.commit()
    et = time.time()
    elapsed_time = et - st
    print('Execution time:', elapsed_time, 'seconds')
